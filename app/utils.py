from io import BytesIO

import openpyxl


async def create_xlsx_report(data: dict) -> bytes:
    workbook = openpyxl.Workbook()
    for account, items in data.items():
        if workbook.sheetnames:
            sheet = workbook.create_sheet(title=account)
        else:
            sheet = workbook.active
            sheet.title = account

        headers = ["Название объявления", "Отвеченные звонки", "Звонки всего", "Новые звонки", "Новые и отвеченные звонки"]
        sheet.append(headers)

        for item in items:
            sheet.append([
                item['title'],
                item['answered'],
                item['calls'],
                item['new'],
                item['newAnswered']
            ])

        for col in sheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    max_length = len(str(cell.value)) if len(str(cell.value)) > max_length else max_length
                except:
                    pass
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[column].width = adjusted_width

    file_stream = BytesIO()
    workbook.save(file_stream)
    file_stream.seek(0)
    return file_stream.read()